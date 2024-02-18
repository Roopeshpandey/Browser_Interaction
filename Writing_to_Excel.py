from selenium import webdriver
import openpyxl
import os

file = 'C:\\Users\\Roope\\OneDrive\\Desktop\\Study_Matterial\\Python_Projects\\Selenium Mini Projects\\Browser_Interaction\\TestData1.xlsx'

# file<<<sheet<<<rows<<<cells

workbook = openpyxl.load_workbook(file)  # load workbook
sheet = workbook.active                  # switch to working sheet
# sheet.cell(r,c).value                  # to get cell value


# when we have to write same data
# for r in range (1,6):
#     for c in range (1,4):
#         sheet.cell(r,c).value='welcome'


sheet.cell(1,1).value = 'A'
sheet.cell(1,2).value = 'B'
sheet.cell(1,3).value = 'C'

sheet.cell(2,1).value = '1'
sheet.cell(2,2).value = '2'
sheet.cell(2,3).value = '3'

sheet.cell(3,1).value = '4'
sheet.cell(3,2).value = '5'
sheet.cell(3,3).value = '6'


workbook.save(file)