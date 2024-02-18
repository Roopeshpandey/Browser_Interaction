from selenium import webdriver
import openpyxl

path = r"C:\Users\Roope\OneDrive\Desktop\Study_Matterial\Python_Projects\Selenium Mini Projects\Browser_Interaction\Test_Excel.xlsx"

# file<<<sheet<<<rows<<<cells

workbook = openpyxl.load_workbook(path)  # load workbook
sheet = workbook["Sheet1"]               # switch to working sheet
# sheet.cell(r,c).value                  # to get cell value

rows = sheet.max_row
cols = sheet.max_column

for r in range (1,rows+1):
    for c in range (1,cols+1):
        print(sheet.cell(r,c).value,end = '      ')
    print()

